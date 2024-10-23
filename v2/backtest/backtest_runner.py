import sys
import os
import pandas as pd
from openpyxl import load_workbook  # openpyxl의 load_workbook 임포트
from openpyxl.utils import get_column_letter

# 프로젝트 루트 경로를 sys.path에 추가
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.abspath(os.path.join(current_dir, '../../'))
sys.path.append(project_root)

from v2.apis.yfinance_history import save_ticker_data_to_csv
from v2.backtest.sma_backtest import sma_backtest_batch_run
from v2.backtest.goden_cross_backtest import golden_backtest_batch_run
from v2.config.constants import RESULT_ANALYSIS_PATH, RESULT_BACKTEST_PATH, RESULT_PERFORMANCE_PATH, TICKERS
from v2.backtest.macd_backtest import macd_backtest_batch_run
from v2.backtest.rsi_backtest import rsi_backtest_batch_run
from v2.backtest.bollinger_backtest import bollinger_backtest_batch_run
from v2.apis.yfinance_stock_name import get_stock_name
from v2.apis.naver_stock_name import get_stock_name_from_naver


COUNT = 365
STRATEGES = ['SMA', 'Golden_Cross', 'MACD', 'RSI', 'Bollinger']

def results_to_file(results_df):
    """
    결과를 파일로 저장
    모바일용 간단한 결과 파일을 저장하는 함수.
    짝수 행에 연한 배경색을 넣어 엑셀 파일로 저장.
    """
    
    # Best strategy 결과 파일로 저장
    csv_file_path = os.path.join(RESULT_ANALYSIS_PATH, 'backtest_best_strategy.csv')
    results_df.to_csv(csv_file_path, index=False)

    # 필요한 열만 추출
    mobile_friendly_df = results_df[['Ticker', 'Return (%)', 'Win Rate (%)', 'Last Action', 'Last Time', 'Last Price', 'KOR Name']].copy()

    # Ticker가 '.KS' 또는 '.KD'를 포함하는 경우 KOR Name 데이터를 Ticker에 넣고 KOR Name 컬럼 삭제
    mobile_friendly_df.loc[mobile_friendly_df['Ticker'].str.contains('.KS') | mobile_friendly_df['Ticker'].str.contains('.KD'), 'Ticker'] = mobile_friendly_df['KOR Name']
    
    # KOR Name 컬럼 삭제
    mobile_friendly_df.drop(columns=['KOR Name'], inplace=True)

    # Return (%), Win Rate (%)는 소수점 없이, Last Price는 소수점 5자리로 포맷팅하고 세 자리마다 쉼표 추가
    mobile_friendly_df['Return (%)'] = mobile_friendly_df['Return (%)'].round(0).astype(int)
    mobile_friendly_df['Win Rate (%)'] = mobile_friendly_df['Win Rate (%)'].round(0).astype(int)
    mobile_friendly_df['Last Price'] = mobile_friendly_df['Last Price'].apply(lambda x: f"{x:,.1f}")

    # 모바일용 결과를 엑셀 파일로 저장
    mobile_xlsx_file_path = os.path.join(RESULT_ANALYSIS_PATH, 'backtest_best_strategy_mobile.xlsx')

    # Styler를 사용하여 짝수 행에 연한 배경색(#ebffeb)을 넣음
    def highlight_even_rows(row):
        return ['background-color: #ebffeb' if row.name % 2 == 0 else '' for _ in row]


    # 'Last Price' 열은 오른쪽 정렬, 'Last Action'과 'Last Time'은 가운데 정렬
    styled_df = mobile_friendly_df.style.apply(highlight_even_rows, axis=1)\
                                       .set_properties(subset=['Last Price'], **{'text-align': 'right'})\
                                       .set_properties(subset=['Last Action', 'Last Time'], **{'text-align': 'center'})

    styled_df.to_excel(mobile_xlsx_file_path, index=False, engine='openpyxl')

    # 열 너비와 행 높이 조정
    adjust_excel_formatting(mobile_xlsx_file_path)

    print(f"결과가 '{csv_file_path}, {mobile_xlsx_file_path}'에 저장되었습니다.")


def adjust_excel_formatting(file_path):
    """
    엑셀 파일의 열 너비와 행 높이를 조정하는 함수
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # 열 너비 조정
    column_widths = {
        'A': 25,  # Ticker
        'B': 12,  # Return (%)
        'C': 12,  # Win Rate (%)
        'D': 12,  # Last Action
        'E': 15,  # Last Time
        'F': 15,  # Last Price
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # 행 높이 조정
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 20  # 모든 행의 높이를 20으로 설정

    # 저장
    workbook.save(file_path)



def evaluation_to_file(content, filename):
    """파일로 저장하는 함수"""
    output_file = os.path.join(RESULT_ANALYSIS_PATH, filename)
    with open(output_file, 'w') as f:
        f.write(content)


def fetch_ticker_data(tickers, count=1000):
    """티커 데이터를 저장합니다."""
    for ticker in tickers:
        save_ticker_data_to_csv(ticker, count)


def run_backtests(tickers, count=100, capital=10000, windows=[5, 20, 60, 120], window_combinations=[(5, 20), (5, 40), (10, 20), (10, 40)], fee=0.001):
    """SMA와 Golden Cross 백테스트를 실행합니다."""

    sma_backtest_batch_run(tickers, count, capital, windows, fee)
    
    golden_backtest_batch_run(tickers, count, capital, window_combinations, fee)
    
    windows=[(12, 16, 9)]
    macd_backtest_batch_run(tickers, count, capital, windows, fee)

    windows=[6, 12,  14, 18, 24, 30]
    rsi_backtest_batch_run(tickers, count, capital, windows=windows, fee=fee)

    windows = [20, 30, 60]
    bollinger_backtest_batch_run(tickers, count, capital, windows=windows, fee=fee)


def evaluate_backtest_results(tickers):
    """백테스트 결과를 평가하고, 최고 전략과 마지막 시그널을 찾아 출력 및 저장합니다."""
    total_string = ''
    results_list = []

    for ticker in tickers:
        ticker_returns = []  # 각 ticker의 전략별 성과를 저장할 리스트

        ticker_header = f"\n\n* {ticker} _____________"
        total_string += ticker_header

        for strategy in STRATEGES:
            strategy_header = f"\n| {strategy}\n"
            total_string += strategy_header

            # 성과 파일 로드
            performance_file = f'{RESULT_PERFORMANCE_PATH}/{strategy}/performance_{strategy}_{ticker}.csv'
            performance_df = pd.read_csv(performance_file)
            df_string = performance_df.to_string(index=False, justify='center', col_space=15) + '\n'
            total_string += df_string

            # 'Return (%)'의 최대값과 그에 해당하는 'Window' 값 찾기
            max_return_idx = performance_df['Return (%)'].idxmax()  # 최대값의 인덱스
            max_return = performance_df['Return (%)'].iloc[max_return_idx]
            max_window = performance_df['Window'].iloc[max_return_idx]
            max_win_rate = performance_df['Win Rate (%)'].iloc[max_return_idx]
            max_mdd = performance_df['Max Drawdown (%)'].iloc[max_return_idx]

            # 해당 ticker의 성과 저장
            ticker_returns.append({
                'Strategy': strategy,
                'Window': max_window,
                'Return': max_return,
                'Win Rate': max_win_rate,
                'MDD': max_mdd
            })

        max_return_strategy = max(ticker_returns, key=lambda x: x['Return'])
        best_strategy = max_return_strategy['Strategy']
        best_window = max_return_strategy['Window']
        best_win_rate = max_return_strategy['Win Rate']
        best_mdd = max_return_strategy['MDD']

        # 각 티커에서 최대 Return을 가진 전략 및 Window 찾기
        # backtest 파일 경로 설정
        backtest_path = os.path.join(RESULT_BACKTEST_PATH, f'{best_strategy}/{best_strategy}_{ticker}/backtest_{best_strategy}_{ticker}_{best_window}.csv')
        backtest_df = pd.read_csv(backtest_path)

        # 마지막 행의 'Action', 'Time', 'Close' 값 가져오기
        last_row = backtest_df.iloc[-1]  # 마지막 행
        last_action = last_row['Action']
        last_time = pd.to_datetime(last_row['Time']).date()
        last_close = last_row['Close']

        kor_stock_name = ''
        if '.KS' in ticker or '.KD' in ticker:
            kor_stock_name = get_stock_name_from_naver(ticker.split(".")[0])

        # Best strategy 저장
        results_list.append({
            'Ticker': ticker,
            'Best Strategy': best_strategy,
            'Window': best_window,
            'Return (%)': max_return_strategy['Return'],
            'Win Rate (%)': best_win_rate, 
            'MDD (%)': best_mdd, 
            'Last Action': last_action,
            'Last Time': last_time,
            'Last Price': last_close,
            'KOR Name': kor_stock_name
        })

    # 백테스트 평가 결과 파일 저장 및 출력
    evaluation_to_file(total_string, 'backtest_evaluation.txt')
    print(total_string)

    # Best strategy DataFrame 생성 및 출력
    results_df = pd.DataFrame(results_list)
    print(f"\n{results_df}")

    # Best strategy 파일로 저장
    results_to_file(results_df)


def main_backtest_process(tickers):
    """백테스트 실행 및 평가 메인 함수"""
    # fetch_ticker_data(tickers)  # 티커 데이터 저장
    run_backtests(tickers, count=COUNT)  # 백테스트 실행
    evaluate_backtest_results(tickers)  # 백테스트 평가


if __name__ == "__main__":
    tickers = TICKERS
    main_backtest_process(tickers)
